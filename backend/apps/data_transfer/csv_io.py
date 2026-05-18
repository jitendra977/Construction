"""
data_transfer.csv_io
─────────────────────
CSV / Excel import-export helpers for the three most commonly bulk-edited
data types: workforce members, materials/inventory, and attendance records.

Export functions return an HttpResponse (CSV or XLSX).
Import functions return a (preview_rows, error_list) tuple for dry-run,
or commit the rows and return (imported_count, skipped_count, error_list).
"""

import csv
import io
import datetime
from typing import Optional

from django.http import HttpResponse
from django.utils import timezone


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _csv_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _xlsx_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fall back to CSV if openpyxl not available
        return _csv_response(filename.replace('.xlsx', '.csv'), headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Header row styling
    hdr_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _read_upload(file_obj) -> tuple[list[str], list[list]]:
    """
    Read an uploaded CSV or XLSX file.
    Returns (headers, rows) where rows is a list of lists.
    Raises ValueError on unsupported format or parse error.
    """
    name = getattr(file_obj, 'name', '').lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            import openpyxl
        except ImportError:
            raise ValueError('openpyxl is required to read Excel files. Use CSV instead.')
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(c.value or '').strip() for c in row] for row in ws.iter_rows()]
        if not all_rows:
            raise ValueError('Empty spreadsheet.')
        headers = all_rows[0]
        rows = all_rows[1:]
    elif name.endswith('.csv') or True:   # default to CSV
        try:
            content = file_obj.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_obj.seek(0)
            content = file_obj.read().decode('latin-1')
        reader = csv.reader(io.StringIO(content))
        all_rows = list(reader)
        if not all_rows:
            raise ValueError('Empty file.')
        headers = [h.strip() for h in all_rows[0]]
        rows = [[c.strip() for c in r] for r in all_rows[1:] if any(c.strip() for c in r)]
    return headers, rows


def _to_date(val: str) -> Optional[datetime.date]:
    """Try common date formats; return None if unparseable."""
    if not val:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Workforce Members
# ─────────────────────────────────────────────────────────────────────────────

WORKFORCE_EXPORT_HEADERS = [
    'id', 'first_name', 'last_name', 'worker_type', 'role',
    'phone', 'email', 'gender', 'nationality',
    'join_date', 'end_date', 'status',
]

WORKFORCE_IMPORT_HEADERS = [
    'first_name', 'last_name', 'worker_type', 'role',
    'phone', 'email', 'gender', 'nationality',
    'join_date', 'end_date',
]


def export_workforce(project_id: int, fmt: str = 'csv') -> HttpResponse:
    from apps.workforce.models import WorkforceMember
    members = (
        WorkforceMember.objects
        .filter(current_project_id=project_id)
        .select_related('role')
        .order_by('_first_name', '_last_name')
    )
    rows = []
    for m in members:
        rows.append([
            str(m.id),
            m.first_name or '',
            m.last_name or '',
            m.worker_type,
            m.role.title if m.role else '',
            m.phone or '',
            m.email or '',
            m.gender or '',
            m.nationality or '',
            str(m.join_date or ''),
            str(m.end_date or ''),
            m.status,
        ])

    ts = datetime.datetime.utcnow().strftime('%Y%m%d')
    if fmt == 'xlsx':
        return _xlsx_response(f'workforce_{project_id}_{ts}.xlsx', WORKFORCE_EXPORT_HEADERS, rows)
    return _csv_response(f'workforce_{project_id}_{ts}.csv', WORKFORCE_EXPORT_HEADERS, rows)


def _parse_workforce_row(idx: int, headers: list[str], row: list[str]) -> tuple[dict, Optional[str]]:
    """Map a CSV/XLSX row to a dict ready for WorkforceMember creation. Returns (data, error)."""
    col = {h.lower().strip(): i for i, h in enumerate(headers)}

    def get(field, default=''):
        i = col.get(field)
        return row[i].strip() if i is not None and i < len(row) else default

    first_name = get('first_name')
    last_name  = get('last_name')
    if not first_name:
        return {}, f'Row {idx}: first_name is required'

    VALID_WORKER_TYPES = {c[0] for c in [
        ('LABOUR',), ('STAFF',), ('SUBCONTRACTOR',), ('FREELANCE',),
        # legacy aliases accepted from older templates
        ('ENGINEER',), ('SUPERVISOR',), ('MANAGER',), ('ADMIN',),
        ('CONTRACTOR',), ('OTHER',),
    ]}
    worker_type_raw = get('worker_type', 'LABOUR').upper()
    # Map legacy aliases → valid WorkerType choices
    _ALIAS = {
        'ENGINEER': 'STAFF', 'SUPERVISOR': 'STAFF', 'MANAGER': 'STAFF',
        'ADMIN': 'STAFF', 'CONTRACTOR': 'SUBCONTRACTOR', 'OTHER': 'LABOUR',
    }
    worker_type = _ALIAS.get(worker_type_raw, worker_type_raw)
    if worker_type not in {'LABOUR', 'STAFF', 'SUBCONTRACTOR', 'FREELANCE'}:
        worker_type = 'LABOUR'

    join_date = _to_date(get('join_date'))
    if not join_date:
        join_date = datetime.date.today()   # required field — default to today

    end_date = _to_date(get('end_date'))

    return {
        'first_name':  first_name,
        'last_name':   last_name,
        'worker_type': worker_type,
        'role_title':  get('role'),       # resolved to FK later
        'phone':       get('phone') or None,
        'email':       get('email') or None,
        'gender':      get('gender') or '',
        'nationality': get('nationality') or '',
        'join_date':   join_date,
        'end_date':    end_date,
    }, None


def dry_run_workforce(file_obj, project_id: int) -> tuple[list[dict], list[dict]]:
    """
    Parse without writing. Returns (preview_rows, errors).
    preview_rows each have: row, first_name, last_name, worker_type, role, action.
    """
    from apps.workforce.models import WorkforceMember
    try:
        headers, rows = _read_upload(file_obj)
    except ValueError as e:
        return [], [{'row': 0, 'message': str(e)}]

    # Dedup key: (first_name, last_name) normalised to lower
    existing = set(
        WorkforceMember.objects
        .filter(current_project_id=project_id)
        .values_list('_first_name', '_last_name')
    )
    existing_keys = {(fn.lower(), ln.lower()) for fn, ln in existing}

    preview, errors = [], []
    for idx, row in enumerate(rows, 2):
        data, err = _parse_workforce_row(idx, headers, row)
        if err:
            errors.append({'row': idx, 'message': err})
            preview.append({'row': idx, 'first_name': '—', 'last_name': '—',
                            'worker_type': '—', 'role': '—', 'action': 'error', 'error': err})
            continue
        key = (data['first_name'].lower(), data['last_name'].lower())
        action = 'exists' if key in existing_keys else 'new'
        preview.append({
            'row':         idx,
            'first_name':  data['first_name'],
            'last_name':   data['last_name'],
            'worker_type': data['worker_type'],
            'role':        data['role_title'],
            'action':      action,
        })

    return preview, errors


def import_workforce(file_obj, project_id: int, created_by, job) -> tuple[int, int, list[dict]]:
    """
    Commit workforce import. Returns (imported, skipped, errors).
    Skips rows whose (first_name, last_name) already exist in this project.
    """
    from apps.workforce.models import WorkforceMember, WorkforceRole
    try:
        headers, rows = _read_upload(file_obj)
    except ValueError as e:
        job.mark_failed(str(e))
        return 0, 0, [{'row': 0, 'message': str(e)}]

    existing = set(
        WorkforceMember.objects
        .filter(current_project_id=project_id)
        .values_list('_first_name', '_last_name')
    )
    existing_keys = {(fn.lower(), ln.lower()) for fn, ln in existing}

    # Build role lookup once (by title, case-insensitive)
    role_map = {r.title.lower(): r for r in WorkforceRole.objects.all()}

    imported, skipped, errors = 0, 0, []

    for idx, row in enumerate(rows, 2):
        data, err = _parse_workforce_row(idx, headers, row)
        if err:
            errors.append({'row': idx, 'message': err})
            skipped += 1
            continue
        key = (data['first_name'].lower(), data['last_name'].lower())
        if key in existing_keys:
            skipped += 1
            continue

        role = role_map.get((data.pop('role_title') or '').lower())
        try:
            # Must call save() individually — bulk_create bypasses save() and
            # therefore skips the auto employee_id generation logic.
            member = WorkforceMember(
                current_project_id = project_id,
                _first_name        = data['first_name'],
                _last_name         = data['last_name'],
                worker_type        = data['worker_type'],
                role               = role,
                _phone             = data['phone'] or '',
                _email             = data['email'] or '',
                gender             = data['gender'],
                nationality        = data['nationality'],
                join_date          = data['join_date'],
                end_date           = data['end_date'],
                status             = 'ACTIVE',
                created_by         = created_by,
            )
            member.save()
            existing_keys.add(key)
            imported += 1
        except Exception as e:
            errors.append({'row': idx, 'message': f'Save failed: {e}'})
            skipped += 1

    job.mark_done(imported, skipped, len(errors), errors)
    return imported, skipped, errors


# ─────────────────────────────────────────────────────────────────────────────
# Materials / Inventory
# ─────────────────────────────────────────────────────────────────────────────

MATERIAL_EXPORT_HEADERS = [
    'id', 'name', 'category', 'unit', 'unit_price',
    'quantity_in_stock', 'reorder_level', 'supplier',
]

MATERIAL_IMPORT_HEADERS = [
    'name', 'category', 'unit', 'unit_price',
    'quantity_in_stock', 'reorder_level', 'supplier',
]


def export_materials(project_id: int, fmt: str = 'csv') -> HttpResponse:
    from apps.resource.models import Material
    materials = (
        Material.objects
        .filter(project_id=project_id)
        .select_related('supplier')
        .order_by('name')
    )
    rows = []
    for m in materials:
        rows.append([
            m.id,
            m.name,
            m.category or '',
            m.unit or '',
            str(m.unit_price or ''),
            str(m.quantity_in_stock or 0),
            str(m.reorder_level or 0),
            m.supplier.name if m.supplier else '',
        ])

    ts = datetime.datetime.utcnow().strftime('%Y%m%d')
    if fmt == 'xlsx':
        return _xlsx_response(f'materials_{project_id}_{ts}.xlsx', MATERIAL_EXPORT_HEADERS, rows)
    return _csv_response(f'materials_{project_id}_{ts}.csv', MATERIAL_EXPORT_HEADERS, rows)


def _parse_material_row(idx: int, headers: list[str], row: list[str]) -> tuple[dict, Optional[str]]:
    col = {h.lower().strip(): i for i, h in enumerate(headers)}

    def get(field, default=''):
        i = col.get(field)
        return row[i].strip() if i is not None and i < len(row) else default

    name = get('name')
    if not name:
        return {}, f'Row {idx}: name is required'

    def _float(val, default=0.0):
        try:
            return float(val) if val else default
        except ValueError:
            return default

    return {
        'name':               name,
        'category':           get('category') or '',
        'unit':               get('unit') or 'pcs',
        'unit_price':         _float(get('unit_price')),
        'quantity_in_stock':  _float(get('quantity_in_stock')),
        'reorder_level':      _float(get('reorder_level')),
        'supplier_name':      get('supplier') or '',
    }, None


def dry_run_materials(file_obj, project_id: int) -> tuple[list[dict], list[dict]]:
    from apps.resource.models import Material
    try:
        headers, rows = _read_upload(file_obj)
    except ValueError as e:
        return [], [{'row': 0, 'message': str(e)}]

    existing = set(Material.objects.filter(project_id=project_id).values_list('name', flat=True))
    preview, errors = [], []
    for idx, row in enumerate(rows, 2):
        data, err = _parse_material_row(idx, headers, row)
        if err:
            errors.append({'row': idx, 'message': err})
            preview.append({'row': idx, 'name': '—', 'unit': '—', 'action': 'error', 'error': err})
            continue
        action = 'exists' if data['name'] in existing else 'new'
        preview.append({
            'row':      idx,
            'name':     data['name'],
            'category': data['category'],
            'unit':     data['unit'],
            'price':    data['unit_price'],
            'stock':    data['quantity_in_stock'],
            'action':   action,
        })
    return preview, errors


def import_materials(file_obj, project_id: int, created_by, job) -> tuple[int, int, list[dict]]:
    from apps.resource.models import Material, Supplier
    try:
        headers, rows = _read_upload(file_obj)
    except ValueError as e:
        job.mark_failed(str(e))
        return 0, 0, [{'row': 0, 'message': str(e)}]

    existing = set(Material.objects.filter(project_id=project_id).values_list('name', flat=True))
    supplier_cache = {}

    def _get_supplier(name):
        if not name:
            return None
        if name not in supplier_cache:
            s, _ = Supplier.objects.get_or_create(project_id=project_id, name=name)
            supplier_cache[name] = s
        return supplier_cache[name]

    imported, skipped, errors = 0, 0, []
    to_create = []

    for idx, row in enumerate(rows, 2):
        data, err = _parse_material_row(idx, headers, row)
        if err:
            errors.append({'row': idx, 'message': err})
            skipped += 1
            continue
        if data['name'] in existing:
            skipped += 1
            continue

        supplier = _get_supplier(data.pop('supplier_name'))
        to_create.append(Material(
            project_id        = project_id,
            name              = data['name'],
            category          = data['category'],
            unit              = data['unit'],
            unit_price        = data['unit_price'],
            quantity_in_stock = data['quantity_in_stock'],
            reorder_level     = data['reorder_level'],
            supplier          = supplier,
        ))
        existing.add(data['name'])

    if to_create:
        Material.objects.bulk_create(to_create, ignore_conflicts=True)
        imported = len(to_create)

    job.mark_done(imported, skipped, len(errors), errors)
    return imported, skipped, errors


# ─────────────────────────────────────────────────────────────────────────────
# Attendance Records
# ─────────────────────────────────────────────────────────────────────────────

ATTENDANCE_EXPORT_HEADERS = [
    'worker_id', 'worker_name', 'date', 'status',
    'check_in', 'check_out', 'overtime_hours', 'notes',
]


def export_attendance(project_id: int, date_from=None, date_to=None, fmt: str = 'csv') -> HttpResponse:
    from apps.attendance.models import DailyAttendance, AttendanceWorker

    qs = DailyAttendance.objects.filter(
        project_id=project_id,
    ).select_related('worker').order_by('date', 'worker__name')

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    rows = []
    for rec in qs:
        rows.append([
            rec.worker_id,
            rec.worker.name,
            str(rec.date),
            rec.status,
            str(rec.check_in or ''),
            str(rec.check_out or ''),
            str(rec.overtime_hours or 0),
            rec.notes or '',
        ])

    ts = datetime.datetime.utcnow().strftime('%Y%m%d')
    if fmt == 'xlsx':
        return _xlsx_response(f'attendance_{project_id}_{ts}.xlsx', ATTENDANCE_EXPORT_HEADERS, rows)
    return _csv_response(f'attendance_{project_id}_{ts}.csv', ATTENDANCE_EXPORT_HEADERS, rows)


# ─────────────────────────────────────────────────────────────────────────────
# Template downloads (blank CSV with correct headers + sample row)
# ─────────────────────────────────────────────────────────────────────────────

TEMPLATES = {
    'workforce': {
        'headers': WORKFORCE_IMPORT_HEADERS,
        'sample':  ['Ali', 'Raza', 'LABOUR', 'Mason', '+923001234567', '',
                    'M', 'Pakistani', '2026-01-01', ''],
    },
    'materials': {
        'headers': MATERIAL_IMPORT_HEADERS,
        'sample':  ['Portland Cement', 'Cement', 'bags', '950', '500', '50', 'Maple Traders'],
    },
}


def download_template(import_type: str, fmt: str = 'csv') -> HttpResponse:
    tpl = TEMPLATES.get(import_type)
    if not tpl:
        raise ValueError(f'No template for type: {import_type}')
    if fmt == 'xlsx':
        return _xlsx_response(
            f'template_{import_type}.xlsx',
            tpl['headers'],
            [tpl['sample']],
        )
    return _csv_response(
        f'template_{import_type}.csv',
        tpl['headers'],
        [tpl['sample']],
    )
